import pands as pd 
from openpyxl import load_workbook
from openpyxl.styles import(PatternFill,Font,Aligment,Border,Side)
from openpyxl.utils import get_column_leter
from date time import datetime

#carregamento de dados

df = pd.read_csv("vendas.csv")
#coluna data
df["data"]=pd.to_datetime(df["data"])
#coluna total
df["total"] = df["quantidade"] *  df["preco_unitario"]
#coluna mes
df["mes"] = df["data"].dt.to_period("M").astype(str)

print("Dados carregados com sucesso")
print(f"   Total de registros: {len(df)}")
print(f"   Periodo: {df['data'].min().date()} até {df['data'].max().date()}")
print()

#Analise 

# Para cada vendedor: soma o total vendido e conta quantas vendas fez
vendas_por_vendedor = (
    df.groupby("vendedor")
    .agg(
        total_vendido=("total", "sum"),      # soma de todos os totais
        qtd_vendas=("total", "count"),       # quantidade de registros
        ticket_medio=("total", "mean")       # média por venda
    )
    .reset_index()                           # transforma o índice em coluna normal
    .sort_values("total_vendido", ascending=False)  # ordena do maior pro menor
)
# --- Vendas por Categoria ---
vendas_por_categoria = (
    df.groupby("categoria")
    .agg(
        total_vendido=("total", "sum"),
        qtd_itens=("quantidade", "sum")
    )
    .reset_index()
    .sort_values("total_vendido", ascending=False)
)
 
# --- Vendas por Mes ---
vendas_por_mes = (
    df.groupby("mes")
    .agg(total_vendido=("total", "sum"))
    .reset_index()
)
 
# --- Vendas por Regiao ---
vendas_por_regiao = (
    df.groupby("regiao")
    .agg(total_vendido=("total", "sum"))
    .reset_index()
    .sort_values("total_vendido", ascending=False)
)
 
# --- Resumo Geral ---
total_geral      = df["total"].sum()
melhor_vendedor  = vendas_por_vendedor.iloc[0]["vendedor"]   # primeiro após ordenar
melhor_categoria = vendas_por_categoria.iloc[0]["categoria"]
melhor_regiao    = vendas_por_regiao.iloc[0]["regiao"]
 
print("Analise concluida:")
print(f"   Faturamento total: R$ {total_geral:,.2f}")
print(f"   Melhor vendedor:   {melhor_vendedor}")
print(f"   Melhor categoria:  {melhor_categoria}")
print()
 
#Criar excel
arquivo_saida = "relatorio_vendas.xlsx"
 
with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
    #criar abas diferentes
    df.to_excel(writer, sheet_name="Dados Brutos", index=False)
 
    vendas_por_vendedor.to_excel(
        writer, sheet_name="Por Vendedor", index=False
    )
    vendas_por_categoria.to_excel(
        writer, sheet_name="Por Categoria", index=False
    )
    vendas_por_mes.to_excel(
        writer, sheet_name="Por Mes", index=False
    )
    vendas_por_regiao.to_excel(
        writer, sheet_name="Por Regiao", index=False
    )
 
print("Arquivo Excel criado. Aplicando formatacao...")
#Formatar Excel
wb = load_workbook(arquivo_saida)   # abre o arquivo que acabamos de criar

 
# Cor amarelo-ouro para cabecalhos
COR_CABECALHO = "1F4E79"    # azul escuro
COR_LINHA_PAR = "D6E4F0"    # azul bem claro
 
fonte_cabecalho = Font(bold=True, color="FFFFFF", size=11)
fonte_normal    = Font(size=10)
alinhamento_centro = Alignment(horizontal="center", vertical="center")
 
borda_fina = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
 
def formatar_aba(ws):
    """Recebe uma aba e aplica cabecalho colorido + linhas alternadas."""
 
    # Cabecalho (linha 1)
    for cell in ws[1]:
        cell.fill      = PatternFill("solid", fgColor=COR_CABECALHO)
        cell.font      = fonte_cabecalho
        cell.alignment = alinhamento_centro
        cell.border    = borda_fina
 
    # Linhas de dados
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        cor = COR_LINHA_PAR if row_idx % 2 == 0 else "FFFFFF"
        for cell in row:
            cell.fill      = PatternFill("solid", fgColor=cor)
            cell.font      = fonte_normal
            cell.alignment = alinhamento_centro
            cell.border    = borda_fina
 
    # Ajustar largura das colunas automaticamente
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_len + 4
 
 
# Aplicar formatacao em todas as abas
for nome_aba in wb.sheetnames:
    formatar_aba(wb[nome_aba])
 
#Resumo executivo
resumo = wb.create_sheet("Resumo", 0)   # cria na primeira posicao
 
# Titulo
resumo["B2"] = "RELATÓRIO DE VENDAS — ANALISE TRIMESTRAL"
resumo["B2"].font      = Font(bold=True, size=14, color="1F4E79")
resumo["B2"].alignment = alinhamento_centro
 
resumo["B3"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
resumo["B3"].font      = Font(italic=True, size=10, color="888888")
resumo["B3"].alignment = alinhamento_centro
 
# Blocos de KPI (indicadores principais)
kpis = [
        ("Faturamento Total",  f"R$ {total_geral:,.2f}"),
        ("Total de Vendas",    str(len(df))),
        ("Melhor Vendedor",    melhor_vendedor),
        ("Top Categoria",      melhor_categoria),
        ("Melhor Regiao",      melhor_regiao),
]
 
linha = 5
for titulo, valor in kpis:
    # Célula do título
    cel_titulo        = resumo.cell(row=linha, column=2, value=titulo)
    cel_titulo.fill   = PatternFill("solid", fgColor=COR_CABECALHO)
    cel_titulo.font   = Font(bold=True, color="FFFFFF", size=10)
    cel_titulo.alignment = alinhamento_centro
    cel_titulo.border = borda_fina
 
    # Célula do valor
    cel_valor         = resumo.cell(row=linha, column=3, value=valor)
    cel_valor.font    = Font(bold=True, size=11, color="1F4E79")
    cel_valor.alignment = alinhamento_centro
    cel_valor.border  = borda_fina
 
    linha += 1
 
# Ajustar largura das colunas do resumo
resumo.column_dimensions["B"].width = 22
resumo.column_dimensions["C"].width = 28
 
 
# ============================================================
# SALVAR TUDO
# ============================================================
 
wb.save(arquivo_saida)
 
print(f"Relatório completo salvo em: {arquivo_saida}")
print()
print("Abas criadas no Excel:")
for aba in wb.sheetnames:
    print(f"   • {aba}")
